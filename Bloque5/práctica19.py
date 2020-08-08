"""En un centro educativo, los datos y calificaciones del alumnado se guarda en un fichero de nombre 
calificaciones.xlsx. La primera columna contiene los apellidos y nombres de los alumnos en orden alfabético, 
las dos siguientes columnas las notas de los dos parciales teóricos realizados  y la última columna 
las notas de prácticas. 
Sabiendo que el peso de cada parcial teórico en la nota final es del 30% y las notas de prácticas cuentan un 40%, 
desarrolla una función que a partir del fichero de calificaciones, devuelva un diccionario donde las claves 
son los apellidos y nombre del alumnado y los valores la nota final calculada."""

import openpyxl

def calcula_notas(fichero):

    excel = openpyxl.load_workbook(fichero)
    hoja = excel.active

    filas = hoja.max_row

    notas = {}

    for fila in range(1,filas+1):
        nombre = hoja.cell(row=fila,column=1).value
        nota_teorica1 = hoja.cell(row=fila,column=2).value
        nota_teorica2 = hoja.cell(row=fila,column=3).value
        nota_practica = hoja.cell(row=fila,column=4).value

        nota_final = nota_teorica1 * 0.3 + nota_teorica2 * 0.3 + nota_practica * 0.4
        notas[nombre] = round(nota_final,2)
    
    return notas












